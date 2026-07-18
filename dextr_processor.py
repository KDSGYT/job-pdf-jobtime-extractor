"""Workbook processing engine for DEXTR Crew Not On Duty updates.

The engine intentionally edits only a copy of the target workbook. Source values
only fill empty cells in columns B through F on the "Crew Not On Duty Report"
worksheet; generated HRs Left formulas are added to column H.
"""

from __future__ import annotations

import re
import shutil
from copy import copy
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

REPORT_SHEET_NAME = "Crew Not On Duty Report"
IGNORED_SHEET_NAME = "Source"
STATUS_CODES = ("DNC", "STB", "WSIB", "LD")
TRANSFER_COLUMNS = ("B", "C", "D", "E", "F")
TRANSFER_COLUMN_NUMBERS = {"B": 2, "C": 3, "D": 4, "E": 5, "F": 6}
PREDICTED_HOURS_COLUMN = 7
HRS_LEFT_COLUMN = 8
HRS_LEFT_HEADING = "HRs Left"
ZERO_PREDICTED_HOURS_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")
GREY_OUT_EXCEPTIONS = {"easter, douglas"}
NEVER_COPY_COLUMNS = ("A", "G", "H")
COLUMN_HEADINGS = {
    "B": "60 Hours",
    "C": "Canvassing Day Shift",
    "D": "Day Shift Notes",
    "E": "Canvassing Night Shift",
    "F": "Night Shift Notes",
}
SECTION_HEADER_NAMES = {"qcto", "cto", "csa", "trainee"}
TITLE_MARKERS = ("crew not on duty report", "employee name", "60 hours")

CLASSIFICATION_PATTERNS = [
    r"\bQCTO\b",
    r"\bCTO\b",
    r"\bCSA\b",
    r"\bSTO\b",
    r"\bTT\b",
    r"\bINACTIVE\b",
    r"\bCrew\s+Dispatcher\s*\d*\b",
    r"\bSupervisor\b",
    r"\bManager\b",
    r"\bGO\s+Transit\b",
    r"\bsick\b",
    r"\btrainee\b",
    r"\bTrainee\b",
]
CLASSIFICATION_RE = re.compile("|".join(f"(?:{p})" for p in CLASSIFICATION_PATTERNS), re.IGNORECASE)
STATUS_RE = re.compile(r"\b(DNC|STB|WSIB|LD)\b", re.IGNORECASE)
REMOVED_TARGET_NAME_RE = re.compile(r"\b(STO|REO|GSR)\b", re.IGNORECASE)
EXCLUDED_SOURCE_VALUE_RE = re.compile(r"\bLVM\s*@", re.IGNORECASE)


@dataclass(frozen=True)
class NameInfo:
    original_value: str
    person_name: str
    job_classification: str


@dataclass(frozen=True)
class SourceValue:
    value: str | None
    column_letter: str
    column_heading: str
    source_label: str
    source_filename: str
    source_worksheet: str
    source_row: int
    status_found: tuple[str, ...] = ()


@dataclass
class SourceEmployeeRow:
    original_column_a: str
    person_name: str
    normalized_name: str
    job_classification: str
    worksheet_section: str | None
    source_label: str
    source_filename: str
    source_worksheet: str
    source_row: int
    values_by_column: dict[str, str]
    statuses_by_column: dict[str, list[str]]


@dataclass
class EmployeeUpdate:
    person_name: str
    normalized_name: str
    original_column_a_values: list[str] = field(default_factory=list)
    rows: list[SourceEmployeeRow] = field(default_factory=list)
    proposed_values: dict[str, SourceValue] = field(default_factory=dict)
    history_by_column: dict[str, list[SourceValue]] = field(default_factory=lambda: {col: [] for col in TRANSFER_COLUMNS})


@dataclass
class PreviewChange:
    change_id: str
    employee_name: str
    normalized_name: str
    target_row: int
    column_letter: str
    column_heading: str
    existing_value: str | None
    proposed_value: str | None
    source_label: str
    source_filename: str
    source_row: int
    status_found: str
    selected: bool = True


@dataclass
class UnmatchedEmployee:
    employee_name: str
    normalized_name: str
    source_files: list[str]
    statuses: list[str]


@dataclass
class ConsolidatedReportRow:
    employee_name: str
    dnc: str
    stb: str
    wsib: str
    ld: str
    column_found: str
    original_cell_value: str
    source_filename: str
    source_worksheet: str
    source_row: int
    oldest_occurrence: str
    most_recent_occurrence: str
    proposed_final_value: str
    match_result: str


def cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_name(name: str) -> str:
    normalized = cell_text(name).lower()
    normalized = re.sub(r"\s*,\s*", ", ", normalized)
    normalized = re.sub(r"[\u2018\u2019]", "'", normalized)
    normalized = re.sub(r"[\u201c\u201d]", '"', normalized)
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized.strip(" .;:-")


def extract_person_name(value: str) -> NameInfo:
    original = cell_text(value)
    working = original
    classifications: list[str] = []

    # Remove parenthesized operational annotations only when the whole note is a
    # known annotation, e.g. "(trainee)". Keep nicknames like "(Sam)".
    def replace_annotation(match: re.Match) -> str:
        inner = match.group(1).strip()
        if re.fullmatch(r"trainee|inactive|sick", inner, flags=re.IGNORECASE):
            classifications.append(inner)
            return " "
        return match.group(0)

    working = re.sub(r"\(([^)]*)\)", replace_annotation, working)

    for match in CLASSIFICATION_RE.finditer(working):
        classifications.append(match.group(0).strip())
    working = CLASSIFICATION_RE.sub(" ", working)
    working = re.sub(r"\s+", " ", working).strip(" -–—,;/")

    return NameInfo(
        original_value=original,
        person_name=working,
        job_classification=", ".join(dict.fromkeys(c for c in classifications if c)),
    )


def find_statuses(value) -> list[str]:
    found = [match.group(1).upper() for match in STATUS_RE.finditer(cell_text(value))]
    return list(dict.fromkeys(found))


def is_excluded_source_value(value) -> bool:
    return bool(EXCLUDED_SOURCE_VALUE_RE.search(cell_text(value)))


def is_ignored_column_a(value) -> bool:
    text = cell_text(value)
    lowered = normalize_name(text)
    if not text:
        return True
    if lowered in SECTION_HEADER_NAMES:
        return True
    if any(marker in lowered for marker in TITLE_MARKERS):
        return True
    # Employee rows in these reports are formatted as "Last, First ...".
    return "," not in text


def worksheet_or_raise(path: Path, sheet_name: str = REPORT_SHEET_NAME) -> Worksheet:
    wb = load_workbook(path, keep_vba=path.suffix.lower() == ".xlsm", data_only=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f'{path.name} does not contain worksheet "{sheet_name}"')
    return wb[sheet_name]


def scan_source_workbook(path: str | Path, source_label: str) -> list[SourceEmployeeRow]:
    workbook_path = Path(path)
    ws = worksheet_or_raise(workbook_path)
    rows: list[SourceEmployeeRow] = []
    current_section: str | None = None

    for row_number in range(1, ws.max_row + 1):
        column_a = cell_text(ws.cell(row=row_number, column=1).value)
        if normalize_name(column_a) in SECTION_HEADER_NAMES:
            current_section = column_a
        if is_ignored_column_a(column_a):
            continue

        statuses_by_column: dict[str, list[str]] = {}
        values_by_column: dict[str, str] = {}
        has_status = False

        for column_letter, column_number in TRANSFER_COLUMN_NUMBERS.items():
            value = cell_text(ws.cell(row=row_number, column=column_number).value)
            if is_excluded_source_value(value):
                continue
            statuses = find_statuses(value)
            if statuses:
                statuses_by_column[column_letter] = statuses
                has_status = True
            if value:
                values_by_column[column_letter] = value

        if not has_status:
            continue

        name_info = extract_person_name(column_a)
        if not name_info.person_name:
            continue

        rows.append(
            SourceEmployeeRow(
                original_column_a=column_a,
                person_name=name_info.person_name,
                normalized_name=normalize_name(name_info.person_name),
                job_classification=name_info.job_classification,
                worksheet_section=current_section,
                source_label=source_label,
                source_filename=workbook_path.name,
                source_worksheet=REPORT_SHEET_NAME,
                source_row=row_number,
                values_by_column=values_by_column,
                statuses_by_column=statuses_by_column,
            )
        )
    return rows


def build_updates_from_sources(
    source_paths: Iterable[str | Path],
    source_labels: Iterable[str] | None = None,
    allow_blank_source_cells_to_clear: bool = False,
) -> dict[str, EmployeeUpdate]:
    paths = list(source_paths)
    labels = list(source_labels or [f"Source {i + 1}" for i in range(len(paths))])
    updates: dict[str, EmployeeUpdate] = {}

    for path, label in zip(paths, labels):
        for row in scan_source_workbook(path, label):
            update = updates.setdefault(
                row.normalized_name,
                EmployeeUpdate(person_name=row.person_name, normalized_name=row.normalized_name),
            )
            update.rows.append(row)
            update.original_column_a_values.append(row.original_column_a)

            for column_letter in TRANSFER_COLUMNS:
                if allow_blank_source_cells_to_clear:
                    value = row.values_by_column.get(column_letter)
                elif column_letter not in row.values_by_column:
                    continue
                else:
                    value = row.values_by_column[column_letter]

                source_value = SourceValue(
                    value=value,
                    column_letter=column_letter,
                    column_heading=COLUMN_HEADINGS[column_letter],
                    source_label=row.source_label,
                    source_filename=row.source_filename,
                    source_worksheet=row.source_worksheet,
                    source_row=row.source_row,
                    status_found=tuple(row.statuses_by_column.get(column_letter, [])),
                )
                update.history_by_column[column_letter].append(source_value)
                if allow_blank_source_cells_to_clear or value not in (None, ""):
                    update.proposed_values[column_letter] = source_value

    return updates


def target_name_index(target_path: str | Path) -> dict[str, tuple[int, str]]:
    ws = worksheet_or_raise(Path(target_path))
    index: dict[str, tuple[int, str]] = {}
    for row_number in range(1, ws.max_row + 1):
        column_a = cell_text(ws.cell(row=row_number, column=1).value)
        if is_ignored_column_a(column_a):
            continue
        name_info = extract_person_name(column_a)
        if name_info.person_name:
            index[normalize_name(name_info.person_name)] = (row_number, name_info.person_name)
    return index


def build_preview_changes(
    target_path: str | Path,
    updates: dict[str, EmployeeUpdate],
) -> tuple[list[PreviewChange], list[UnmatchedEmployee]]:
    target = Path(target_path)
    ws = worksheet_or_raise(target)
    index = target_name_index(target)
    changes: list[PreviewChange] = []
    unmatched: list[UnmatchedEmployee] = []

    for normalized_name, update in sorted(updates.items(), key=lambda item: item[1].person_name.lower()):
        if normalized_name not in index:
            statuses = sorted({status for row in update.rows for found in row.statuses_by_column.values() for status in found})
            unmatched.append(
                UnmatchedEmployee(
                    employee_name=update.person_name,
                    normalized_name=normalized_name,
                    source_files=sorted({row.source_filename for row in update.rows}),
                    statuses=statuses,
                )
            )
            continue

        target_row, display_name = index[normalized_name]
        for column_letter in TRANSFER_COLUMNS:
            if column_letter not in update.proposed_values:
                continue
            source_value = update.proposed_values[column_letter]
            existing = ws.cell(row=target_row, column=TRANSFER_COLUMN_NUMBERS[column_letter]).value
            if cell_text(existing):
                continue
            proposed = source_value.value
            changes.append(
                PreviewChange(
                    change_id=f"{normalized_name}|{target_row}|{column_letter}",
                    employee_name=display_name,
                    normalized_name=normalized_name,
                    target_row=target_row,
                    column_letter=column_letter,
                    column_heading=COLUMN_HEADINGS[column_letter],
                    existing_value=cell_text(existing) or None,
                    proposed_value=proposed,
                    source_label=source_value.source_label,
                    source_filename=source_value.source_filename,
                    source_row=source_value.source_row,
                    status_found=", ".join(source_value.status_found),
                    selected=True,
                )
            )
    return changes, unmatched


def remove_ineligible_target_rows(ws: Worksheet) -> int:
    removed = 0
    for row_number in range(ws.max_row, 0, -1):
        column_a = cell_text(ws.cell(row=row_number, column=1).value)
        if REMOVED_TARGET_NAME_RE.search(column_a):
            ws.delete_rows(row_number)
            removed += 1
    return removed


def add_hrs_left_formulas(ws: Worksheet) -> int:
    formulas_added = 0
    in_focus_table = False
    for row_number in range(1, ws.max_row + 1):
        column_a = cell_text(ws.cell(row=row_number, column=1).value)
        normalized_a = normalize_name(column_a)
        row_has_values = any(
            cell_text(ws.cell(row=row_number, column=col).value)
            for col in range(1, PREDICTED_HOURS_COLUMN + 1)
        )

        if not row_has_values:
            in_focus_table = False
            continue
        if normalized_a in SECTION_HEADER_NAMES:
            in_focus_table = True
            ws.cell(row=row_number, column=HRS_LEFT_COLUMN, value=HRS_LEFT_HEADING)
            continue
        if in_focus_table and column_a:
            ws.cell(row=row_number, column=HRS_LEFT_COLUMN, value=f"=60-G{row_number}")
            formulas_added += 1
    return formulas_added


def predicted_hours_value(ws: Worksheet, row_number: int) -> float | None:
    predicted_hours = ws.cell(row=row_number, column=PREDICTED_HOURS_COLUMN).value
    if isinstance(predicted_hours, (int, float)):
        return float(predicted_hours)
    try:
        return float(cell_text(predicted_hours))
    except ValueError:
        return None


def hrs_left_sort_value(ws: Worksheet, row_number: int) -> tuple[int, float, int]:
    """Sort key for HRs Left, keeping rows without numeric hours at the end."""
    hrs_left = ws.cell(row=row_number, column=HRS_LEFT_COLUMN).value
    if isinstance(hrs_left, (int, float)):
        return (0, float(hrs_left), row_number)

    predicted_hours = predicted_hours_value(ws, row_number)
    if predicted_hours is None:
        return (1, 0, row_number)
    return (0, 60 - predicted_hours, row_number)


def row_snapshot(ws: Worksheet, row_number: int) -> list[dict]:
    snapshot = []
    for column_number in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_number, column=column_number)
        snapshot.append(
            {
                "value": cell.value,
                "style": copy(cell._style),
                "number_format": cell.number_format,
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "border": copy(cell.border),
                "alignment": copy(cell.alignment),
                "protection": copy(cell.protection),
                "comment": copy(cell.comment),
                "hyperlink": copy(cell.hyperlink),
            }
        )
    return snapshot


def restore_row_snapshot(ws: Worksheet, row_number: int, snapshot: list[dict]) -> None:
    for column_number, saved in enumerate(snapshot, start=1):
        cell = ws.cell(row=row_number, column=column_number)
        cell.value = saved["value"]
        cell._style = copy(saved["style"])
        cell.number_format = saved["number_format"]
        cell.font = copy(saved["font"])
        cell.fill = copy(saved["fill"])
        cell.border = copy(saved["border"])
        cell.alignment = copy(saved["alignment"])
        cell.protection = copy(saved["protection"])
        cell.comment = copy(saved["comment"])
        cell.hyperlink = copy(saved["hyperlink"])


def sort_focus_tables_by_hrs_left(ws: Worksheet) -> int:
    """Sort employee rows inside QCTO/CTO/CSA/Trainee tables by HRs Left ascending."""
    rows_reordered = 0
    row_number = 1

    while row_number <= ws.max_row:
        column_a = cell_text(ws.cell(row=row_number, column=1).value)
        if normalize_name(column_a) not in SECTION_HEADER_NAMES:
            row_number += 1
            continue

        first_employee_row = row_number + 1
        last_employee_row = first_employee_row - 1
        scan_row = first_employee_row
        while scan_row <= ws.max_row:
            row_has_values = any(
                cell_text(ws.cell(row=scan_row, column=col).value)
                for col in range(1, HRS_LEFT_COLUMN + 1)
            )
            if not row_has_values:
                break
            if normalize_name(ws.cell(row=scan_row, column=1).value) in SECTION_HEADER_NAMES:
                break
            last_employee_row = scan_row
            scan_row += 1

        if last_employee_row >= first_employee_row:
            row_numbers = list(range(first_employee_row, last_employee_row + 1))
            sorted_row_numbers = sorted(row_numbers, key=lambda item: hrs_left_sort_value(ws, item))
            if sorted_row_numbers != row_numbers:
                snapshots = [row_snapshot(ws, source_row) for source_row in sorted_row_numbers]
                for target_row, snapshot in zip(row_numbers, snapshots):
                    restore_row_snapshot(ws, target_row, snapshot)
                    ws.cell(row=target_row, column=HRS_LEFT_COLUMN, value=f"=60-G{target_row}")
                rows_reordered += len(row_numbers)

        row_number = max(scan_row, row_number + 1)

    return rows_reordered


def row_has_dnc(ws: Worksheet, row_number: int) -> bool:
    return any(
        not is_excluded_source_value(ws.cell(row=row_number, column=col).value)
        and "DNC" in find_statuses(ws.cell(row=row_number, column=col).value)
        for col in TRANSFER_COLUMN_NUMBERS.values()
    )


def is_grey_out_exception(value) -> bool:
    name_info = extract_person_name(cell_text(value))
    return normalize_name(name_info.person_name) in GREY_OUT_EXCEPTIONS


def shade_attention_rows(ws: Worksheet) -> int:
    shaded = 0
    in_focus_table = False
    for row_number in range(1, ws.max_row + 1):
        column_a = cell_text(ws.cell(row=row_number, column=1).value)
        normalized_a = normalize_name(column_a)
        row_has_values = any(
            cell_text(ws.cell(row=row_number, column=col).value)
            for col in range(1, HRS_LEFT_COLUMN + 1)
        )

        if not row_has_values:
            in_focus_table = False
            continue
        if normalized_a in SECTION_HEADER_NAMES:
            in_focus_table = True
            continue
        if is_grey_out_exception(column_a):
            continue
        if in_focus_table and column_a and (predicted_hours_value(ws, row_number) == 0 or row_has_dnc(ws, row_number)):
            for column_number in range(1, ws.max_column + 1):
                ws.cell(row=row_number, column=column_number).fill = copy(ZERO_PREDICTED_HOURS_FILL)
            shaded += 1
    return shaded


def apply_selected_changes(
    target_path: str | Path,
    output_path: str | Path,
    changes: Iterable[PreviewChange],
) -> Path:
    target = Path(target_path)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(target, output)

    wb = load_workbook(output, keep_vba=True, data_only=False)
    ws = wb[REPORT_SHEET_NAME]
    for change in changes:
        if not change.selected:
            continue
        if change.column_letter not in TRANSFER_COLUMNS:
            continue
        ws.cell(
            row=change.target_row,
            column=TRANSFER_COLUMN_NUMBERS[change.column_letter],
            value=change.proposed_value,
        )
    remove_ineligible_target_rows(ws)
    add_hrs_left_formulas(ws)
    sort_focus_tables_by_hrs_left(ws)
    shade_attention_rows(ws)
    wb.save(output)
    return output


def consolidated_report_rows(
    updates: dict[str, EmployeeUpdate],
    matched_names: set[str] | None = None,
) -> list[ConsolidatedReportRow]:
    matched_names = matched_names or set()
    report: list[ConsolidatedReportRow] = []
    for normalized_name, update in sorted(updates.items(), key=lambda item: item[1].person_name.lower()):
        for column_letter in TRANSFER_COLUMNS:
            history = update.history_by_column.get(column_letter, [])
            if not history:
                continue
            proposed = update.proposed_values.get(column_letter)
            for item in history:
                if not item.status_found:
                    continue
                status_flags = {status: "Yes" if status in item.status_found else "" for status in STATUS_CODES}
                report.append(
                    ConsolidatedReportRow(
                        employee_name=update.person_name,
                        dnc=status_flags["DNC"],
                        stb=status_flags["STB"],
                        wsib=status_flags["WSIB"],
                        ld=status_flags["LD"],
                        column_found=column_letter,
                        original_cell_value=item.value or "",
                        source_filename=item.source_filename,
                        source_worksheet=item.source_worksheet,
                        source_row=item.source_row,
                        oldest_occurrence=f"{history[0].source_label} row {history[0].source_row}",
                        most_recent_occurrence=f"{history[-1].source_label} row {history[-1].source_row}",
                        proposed_final_value=(proposed.value if proposed else "") or "",
                        match_result="Matched" if normalized_name in matched_names else "Not found in target",
                    )
                )
    return report
