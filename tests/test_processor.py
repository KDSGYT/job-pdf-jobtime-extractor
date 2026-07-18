from pathlib import Path

from openpyxl import Workbook, load_workbook

from dextr_processor import (
    STATUS_CODES,
    apply_selected_changes,
    build_preview_changes,
    build_updates_from_sources,
    extract_person_name,
    normalize_name,
    scan_source_workbook,
)


SHEET = "Crew Not On Duty Report"


def make_workbook(path: Path, rows):
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = SHEET
    ws.append([
        "Employee", "60 Hours", "Canvassing Day Shift", "Day Shift Notes",
        "Canvassing Night Shift", "Night Shift Notes", "Predicted Hours", "HRs Left"
    ])
    for row in rows:
        ws.append(row)
    source = wb.create_sheet("Source")
    source["A1"] = "do not touch"
    wb.save(path)


def test_extract_person_name_removes_job_classifications_but_keeps_real_parentheses():
    assert extract_person_name("Labelle, Paul QCTO").person_name == "Labelle, Paul"
    assert extract_person_name("Igaal, Warsame (Sam) CSA").person_name == "Igaal, Warsame (Sam)"
    assert extract_person_name("Addison, Aaron CTO (trainee)").person_name == "Addison, Aaron"
    assert extract_person_name("Vargas, (Abreu) Paul CSA").person_name == "Vargas, (Abreu) Paul"


def test_normalize_name_matches_case_spaces_and_comma_spacing():
    assert normalize_name(" Labelle,   Paul ") == normalize_name("labelle , paul")


def test_scan_source_workbook_finds_statuses_only_in_columns_b_to_f(tmp_path):
    source = tmp_path / "source1.xlsx"
    make_workbook(source, [
        ["QCTO", "DNC", None, None, None, None, "DO NOT COPY", "=60-G2"],
        ["Crew Not On Duty Report", None, None, None, None, None, None, None],
        ["Labelle, Paul QCTO", None, "DNC", "call tomorrow", "STB - awaiting update", None, "DO NOT COPY", "=60-G4"],
        ["Applewhite, Jeff CTO", None, None, None, None, None, "DNC in G ignored", "=60-G5"],
        [None, None, None, None, None, None, None, None],
    ])

    rows = scan_source_workbook(source, "Source 1")

    assert len(rows) == 1
    row = rows[0]
    assert row.person_name == "Labelle, Paul"
    assert row.values_by_column == {"C": "DNC", "D": "call tomorrow", "E": "STB - awaiting update"}
    assert row.statuses_by_column == {"C": ["DNC"], "E": ["STB"]}
    assert all(status in STATUS_CODES for status in ["DNC", "STB", "WSIB", "LD"])


def test_scan_source_workbook_ignores_lvm_time_values(tmp_path):
    source = tmp_path / "source_lvm.xlsx"
    make_workbook(source, [
        ["Voicemail, Val QCTO", "LVM@ 0900", None, None, None, None, None, None],
        ["Mixed, Mira QCTO", "DNC", "LVM @ 1015", None, None, None, None, None],
    ])

    rows = scan_source_workbook(source, "Source 1")

    assert [row.person_name for row in rows] == ["Mixed, Mira"]
    assert rows[0].values_by_column == {"B": "DNC"}
    assert rows[0].statuses_by_column == {"B": ["DNC"]}


def test_newest_nonblank_source_value_wins_and_blanks_do_not_clear(tmp_path):
    oldest = tmp_path / "oldest.xlsx"
    newest = tmp_path / "newest.xlsx"
    make_workbook(oldest, [["Labelle, Paul QCTO", None, "DNC", "old note", None, None, None, None]])
    make_workbook(newest, [["Labelle, Paul QCTO", None, "STB", None, "LD", "new night note", None, None]])

    updates = build_updates_from_sources([oldest, newest], ["Source 1", "Source 2"])
    update = updates[normalize_name("Labelle, Paul")]

    assert update.proposed_values["C"].value == "STB"
    assert update.proposed_values["C"].source_label == "Source 2"
    assert update.proposed_values["D"].value == "old note"
    assert update.proposed_values["D"].source_label == "Source 1"
    assert update.proposed_values["E"].value == "LD"
    assert update.proposed_values["F"].value == "new night note"


def test_preview_matches_target_rows_and_only_empty_columns_b_to_f(tmp_path):
    source = tmp_path / "source.xlsx"
    target = tmp_path / "target.xlsm"
    make_workbook(source, [["Labelle, Paul QCTO", "DNC", "STB", "day note", "LD", "night note", "source predicted", "source formula"]])
    make_workbook(target, [["Labelle, Paul CTO", "existing B", None, "existing D", None, None, 44, "=60-G2"]])

    updates = build_updates_from_sources([source], ["Source 1"])
    changes, unmatched = build_preview_changes(target, updates)

    assert unmatched == []
    assert [(change.column_letter, change.proposed_value) for change in changes] == [
        ("C", "STB"),
        ("E", "LD"),
        ("F", "night note"),
    ]
    assert all(change.target_row == 2 for change in changes)


def test_apply_selected_changes_preserves_existing_target_data_g_h_formulas_and_source_sheet(tmp_path):
    source = tmp_path / "source.xlsx"
    target = tmp_path / "target.xlsm"
    output = tmp_path / "updated.xlsm"
    make_workbook(source, [["Labelle, Paul QCTO", "DNC", "STB", "day note", None, None, "source predicted", "source formula"]])
    make_workbook(target, [["Labelle, Paul CTO", "existing B", None, "existing D", None, None, 44, "=60-G2"]])

    updates = build_updates_from_sources([source], ["Source 1"])
    changes, _ = build_preview_changes(target, updates)
    apply_selected_changes(target, output, changes)

    wb = load_workbook(output, keep_vba=True, data_only=False)
    ws = wb[SHEET]
    assert ws["A2"].value == "Labelle, Paul CTO"
    assert ws["B2"].value == "existing B"
    assert ws["C2"].value == "STB"
    assert ws["D2"].value == "existing D"
    assert ws["G2"].value == 44
    assert ws["H2"].value == "=60-G2"
    assert wb["Source"]["A1"].value == "do not touch"


def test_apply_selected_changes_removes_target_rows_with_sto_reo_or_gsr_names_case_insensitive(tmp_path):
    source = tmp_path / "source.xlsx"
    target = tmp_path / "target.xlsm"
    output = tmp_path / "updated.xlsm"
    make_workbook(source, [["Keeper, Kim CTO", None, "DNC", None, None, None, None, None]])
    make_workbook(target, [
        ["Keeper, Kim CTO", None, None, None, None, None, None, None],
        ["Remove, Sam STO", None, None, None, None, None, None, None],
        ["Remove, Ray reo", None, None, None, None, None, None, None],
        ["Remove, Gia GsR", None, None, None, None, None, None, None],
        ["Stay, Pat CTO", None, None, None, None, None, None, None],
    ])

    updates = build_updates_from_sources([source], ["Source 1"])
    changes, _ = build_preview_changes(target, updates)
    apply_selected_changes(target, output, changes)

    ws = load_workbook(output, keep_vba=True, data_only=False)[SHEET]
    remaining_names = [ws.cell(row=row, column=1).value for row in range(2, ws.max_row + 1)]

    assert remaining_names == ["Keeper, Kim CTO", "Stay, Pat CTO"]
    assert ws["C2"].value == "DNC"


def test_apply_selected_changes_adds_hrs_left_formulas_to_focus_tables_only(tmp_path):
    target = tmp_path / "target.xlsx"
    output = tmp_path / "updated.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = SHEET
    ws.append(["Crew Not On Duty"])
    ws.append([])
    ws.append(["QCTO", "60 Hours", "Canvassing Day Shift", "Notes", "Canvassing Night Shift", "Notes", "Predicted Hours"])
    ws.append(["Keep, Quinn QCTO", None, None, None, None, None, 37.4])
    ws.append([])
    ws.append(["GSR", "60 Hours", "Canvassing Day Shift", "Notes", "Canvassing Night Shift", "Notes", "Predicted Hours"])
    ws.append(["Skip, Glen GSR", None, None, None, None, None, 42])
    ws.append([])
    ws.append(["Trainee", "60 Hours", "Canvassing Day Shift", "Notes", "Canvassing Night Shift", "Notes", "Predicted Hours"])
    ws.append(["Keep, Tina TT (trainee)", None, None, None, None, None, 23.5])
    wb.save(target)

    apply_selected_changes(target, output, [])

    ws = load_workbook(output, data_only=False)[SHEET]
    assert ws["H3"].value == "HRs Left"
    assert ws["H4"].value == "=60-G4"
    assert ws["A6"].value is None
    assert ws["H6"].value is None
    assert ws["A7"].value == "Trainee"
    assert ws["H7"].value == "HRs Left"
    assert ws["H8"].value == "=60-G8"


def test_apply_selected_changes_sorts_focus_table_rows_by_hrs_left_smallest_first(tmp_path):
    target = tmp_path / "target.xlsx"
    output = tmp_path / "updated.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = SHEET
    ws.append(["QCTO", "60 Hours", "Canvassing Day Shift", "Notes", "Canvassing Night Shift", "Notes", "Predicted Hours"])
    ws.append(["High, Hours QCTO", None, None, "keep row data", None, None, "58.00"])  # 2 HRs left
    ws.append(["Low, Hours QCTO", None, None, None, None, None, "10.00"])  # 50 HRs left
    ws.append(["Zero, Hours QCTO", None, None, None, None, None, "0.00"])  # 60 HRs left
    ws.append(["Middle, Hours QCTO", None, None, None, None, None, "44.00"])  # 16 HRs left
    ws.append([])
    ws.append(["GSR", "60 Hours", "Canvassing Day Shift", "Notes", "Canvassing Night Shift", "Notes", "Predicted Hours"])
    ws.append(["Unsorted, Glen GSR", None, None, None, None, None, 59])
    wb.save(target)

    apply_selected_changes(target, output, [])

    ws = load_workbook(output, data_only=False)[SHEET]
    assert [ws.cell(row=row, column=1).value for row in range(2, 6)] == [
        "High, Hours QCTO",
        "Middle, Hours QCTO",
        "Low, Hours QCTO",
        "Zero, Hours QCTO",
    ]
    assert [ws.cell(row=row, column=7).value for row in range(2, 6)] == ["58.00", "44.00", "10.00", "0.00"]
    assert [ws.cell(row=row, column=8).value for row in range(2, 6)] == ["=60-G2", "=60-G3", "=60-G4", "=60-G5"]
    assert ws["D2"].value == "keep row data"
    assert ws["A2"].fill.fgColor.rgb != "00D9D9D9"
    assert ws["A5"].fill.fgColor.rgb == "00D9D9D9"
    assert ws["H5"].fill.fgColor.rgb == "00D9D9D9"
    assert ws["A4"].fill.fgColor.rgb != "00D9D9D9"
    assert "GSR" not in [ws.cell(row=row, column=1).value for row in range(1, ws.max_row + 1)]


def test_apply_selected_changes_greys_dnc_but_not_other_notes_or_statuses(tmp_path):
    target = tmp_path / "target.xlsx"
    output = tmp_path / "updated.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = SHEET
    ws.append(["QCTO", "60 Hours", "Canvassing Day Shift", "Notes", "Canvassing Night Shift", "Notes", "Predicted Hours"])
    ws.append(["Dnc, Dana QCTO", "DNC", None, None, None, None, "55.00"])
    ws.append(["Stb, Sam QCTO", None, "STB", None, None, None, "54.00"])
    ws.append(["Note, Nora QCTO", None, None, "call back", None, None, "53.00"])
    ws.append(["Lvm, Liam QCTO", "LVM@ 0900", None, None, None, None, "52.00"])
    ws.append(["Easter, Douglas QCTO", "DNC", None, None, None, None, "0.00"])
    ws.append(["Zero, Zack QCTO", None, None, None, None, None, "0.00"])
    wb.save(target)

    apply_selected_changes(target, output, [])

    ws = load_workbook(output, data_only=False)[SHEET]
    rows = {ws.cell(row=row, column=1).value: row for row in range(2, 8)}
    grey = "00D9D9D9"
    assert ws.cell(row=rows["Dnc, Dana QCTO"], column=1).fill.fgColor.rgb == grey
    assert ws.cell(row=rows["Zero, Zack QCTO"], column=1).fill.fgColor.rgb == grey
    assert ws.cell(row=rows["Stb, Sam QCTO"], column=1).fill.fgColor.rgb != grey
    assert ws.cell(row=rows["Note, Nora QCTO"], column=1).fill.fgColor.rgb != grey
    assert ws.cell(row=rows["Lvm, Liam QCTO"], column=1).fill.fgColor.rgb != grey
    assert ws.cell(row=rows["Easter, Douglas QCTO"], column=1).fill.fgColor.rgb != grey
